from __future__ import annotations

import argparse
from pathlib import Path
from statistics import mean

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from report_common import (
    GREEN,
    LIGHT_BLUE,
    RED,
    TEXT,
    WHITE,
    YELLOW,
    add_bar_chart,
    attendance_mark,
    iter_students,
    load_metrics,
    new_workbook,
    performance_fill,
    prepare_output,
    require_teacher_subject,
    select_groups,
    select_subjects,
    set_widths,
    score_to_100,
    student_display_label,
    student_subject,
    style_data_cell,
    style_headers,
    style_section_header,
    title_row,
)


SCRIPT_DIR = Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Сформировать подробный XLSX-отчёт преподавателя по одному предмету."
    )
    parser.add_argument("--teacher-id", type=int, required=True, help="ID преподавателя")
    parser.add_argument("--subject-id", type=int, required=True, help="ID предмета")
    parser.add_argument("--group-ids", type=int, nargs="*", help="ID групп; без параметра — все группы")
    parser.add_argument(
        "--input",
        default=str(SCRIPT_DIR / "fake_metrics.json"),
        help="Путь к исходному JSON",
    )
    parser.add_argument(
        "--output",
        default=str(SCRIPT_DIR / "output" / "teacher_report.xlsx"),
        help="Путь к XLSX",
    )
    return parser.parse_args()


def build_report(args: argparse.Namespace) -> str:
    data = load_metrics(args.input)
    groups = select_groups(data, args.group_ids)
    subject = select_subjects(data, [args.subject_id])[0]
    require_teacher_subject(subject, args.teacher_id)
    filtered_groups = []
    for group in groups:
        students = [
            student
            for student in group["students"]
            if any(item["subject_id"] == subject["subject_id"] for item in student["subjects"])
        ]
        if students:
            filtered_groups.append({**group, "students": students})
    groups = filtered_groups
    if not groups:
        raise SystemExit("Для выбранного предмета и групп нет данных.")

    workbook = new_workbook(f"Отчёт преподавателя: {subject['name']}")
    report = workbook.create_sheet("Предмет")

    sample = student_subject(groups[0]["students"][0], subject["subject_id"])
    lectures = sample["activities"]["lectures"]
    labs = sample["activities"]["laboratory_works"]
    practices = sample["activities"]["practices"]

    fixed_start = 1
    lecture_start = 4
    lab_start = lecture_start + len(lectures)
    practice_start = lab_start + len(labs)
    summary_start = practice_start + len(practices)
    last_column = summary_start + 1

    title_row(
        report,
        f"{subject['name']} — подробный отчёт",
        last_column,
        f"Преподаватель: {subject['teacher']} | Кафедра: {subject['department_name']}",
    )
    report.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    report.cell(
        3,
        1,
        "Лекции: пусто — присутствовал, П — пропуск, Б — болел со справкой. "
        "Лабораторные и практики показаны в 100-балльной шкале; пусто — оценки пока нет.",
    )
    report.cell(3, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    report.cell(3, 1).font = Font(italic=True, color=TEXT)

    section_row = 5
    header_row = 6
    style_section_header(report, section_row, fixed_start, 3, "Студент")
    style_section_header(report, section_row, lecture_start, lab_start - 1, "Лекции")
    style_section_header(report, section_row, lab_start, practice_start - 1, "Лабораторные")
    style_section_header(report, section_row, practice_start, summary_start - 1, "Практики")
    style_section_header(report, section_row, summary_start, last_column, "Итоги")

    headers = ["№", "ФИО / хеш", "Группа"]
    headers.extend(item["date"] for item in lectures)
    headers.extend(f"ЛР №{index + 1}\n{item['date']}" for index, item in enumerate(labs))
    headers.extend(f"Пр №{index + 1}\n{item['date']}" for index, item in enumerate(practices))
    headers.extend(["Успеваемость, %", "Посещаемость, %"])
    for column, value in enumerate(headers, 1):
        report.cell(header_row, column, value)
    style_headers(report, header_row, 1, last_column)

    row = header_row + 1
    for number, (group, student) in enumerate(iter_students(groups), 1):
        metrics = student_subject(student, subject["subject_id"])
        values = [number, student_display_label(student), group["group_name"]]
        values.extend(attendance_mark(item) for item in metrics["activities"]["lectures"])
        values.extend(score_to_100(item) for item in metrics["activities"]["laboratory_works"])
        values.extend(score_to_100(item) for item in metrics["activities"]["practices"])
        values.extend(
            [
                metrics["assessment_summary"]["performance_percent"],
                metrics["attendance_summary"]["attendance_percent"],
            ]
        )

        for column, value in enumerate(values, 1):
            cell = report.cell(row, column, value)
            style_data_cell(cell, "left" if column == 2 else "center")
            if lecture_start <= column < lab_start:
                if value == "П":
                    cell.fill = PatternFill("solid", fgColor=RED)
                elif value == "Б":
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
            elif lab_start <= column < practice_start and isinstance(value, (int, float)):
                cell.fill = performance_fill(value)
                cell.number_format = "0.##"
            elif practice_start <= column < summary_start:
                cell.fill = PatternFill("solid", fgColor=WHITE if value is None else GREEN)
                if isinstance(value, (int, float)):
                    cell.fill = performance_fill(value)
                    cell.number_format = "0.##"
            elif column >= summary_start:
                cell.fill = performance_fill(value)
                cell.number_format = "0.00"
        row += 1

    report.freeze_panes = "D7"
    report.auto_filter.ref = f"A{header_row}:{report.cell(row - 1, last_column).coordinate}"
    report.sheet_view.showGridLines = False
    report.row_dimensions[3].height = 28
    set_widths(report, {1: 7, 2: 29, 3: 13})
    for column in range(lecture_start, summary_start):
        report.column_dimensions[get_column_letter(column)].width = 12
    set_widths(report, {summary_start: 17, last_column: 18})

    analytics = workbook.create_sheet("Аналитика")
    title_row(
        analytics,
        f"Аналитика по предмету «{subject['name']}»",
        4,
        "Средние показатели выбранных групп",
    )
    analytics.append([])
    analytics.append(["Группа", "Студентов", "Успеваемость, %", "Посещаемость, %"])
    style_headers(analytics, 4, 1, 4)

    for group in groups:
        metrics = [student_subject(student, subject["subject_id"]) for student in group["students"]]
        analytics.append(
            [
                group["group_name"],
                len(metrics),
                round(mean(item["assessment_summary"]["performance_percent"] for item in metrics), 2),
                round(mean(item["attendance_summary"]["attendance_percent"] for item in metrics), 2),
            ]
        )

    for data_row in range(5, 5 + len(groups)):
        for column in range(1, 5):
            style_data_cell(analytics.cell(data_row, column), "left" if column == 1 else "center")
        analytics.cell(data_row, 3).fill = performance_fill(analytics.cell(data_row, 3).value)
        analytics.cell(data_row, 4).fill = performance_fill(analytics.cell(data_row, 4).value)

    max_row = 4 + len(groups)
    add_bar_chart(analytics, "Успеваемость по группам", 3, 1, 5, max_row, "F4")
    add_bar_chart(analytics, "Посещаемость по группам", 4, 1, 5, max_row, "F20")
    set_widths(analytics, {1: 20, 2: 14, 3: 20, 4: 20})
    analytics.sheet_view.showGridLines = False

    output = prepare_output(args.output)
    workbook.save(output)
    return str(output)


if __name__ == "__main__":
    arguments = parse_args()
    path = build_report(arguments)
    print(f"Отчёт преподавателя создан: {path}")
