from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLUE = "1756D8"
DARK_BLUE = "123B92"
LIGHT_BLUE = "EAF1FF"
CYAN = "13DCE0"
GREEN = "C6E0B4"
YELLOW = "FFF45C"
RED = "F4A3A3"
PALE_RED = "FCE8E8"
WHITE = "FFFFFF"
TEXT = "10213F"
MUTED = "667694"
BORDER_COLOR = "B7C7E5"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def load_metrics(path: str | Path) -> dict:
    source = Path(path)
    try:
        with source.open("r", encoding="utf-8") as file:
            data = json.load(file)
    except FileNotFoundError as exc:
        raise SystemExit(f"JSON-файл не найден: {source}") from exc
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Некорректный JSON: {exc}") from exc

    required = {"subjects", "groups", "access_control"}
    missing = required.difference(data)
    if missing:
        raise SystemExit(f"В JSON отсутствуют обязательные поля: {', '.join(sorted(missing))}")
    return data


def select_groups(data: dict, group_ids: Sequence[int] | None) -> list[dict]:
    groups = data["groups"]
    if not group_ids:
        return groups

    requested = set(group_ids)
    selected = [group for group in groups if group["group_id"] in requested]
    missing = requested.difference(group["group_id"] for group in selected)
    if missing:
        raise SystemExit(f"Не найдены группы: {', '.join(map(str, sorted(missing)))}")
    return selected


def select_subjects(data: dict, subject_ids: Sequence[int] | None) -> list[dict]:
    subjects = data["subjects"]
    if not subject_ids:
        return subjects

    requested = set(subject_ids)
    selected = [subject for subject in subjects if subject["subject_id"] in requested]
    missing = requested.difference(subject["subject_id"] for subject in selected)
    if missing:
        raise SystemExit(f"Не найдены предметы: {', '.join(map(str, sorted(missing)))}")
    return selected


def require_teacher_subject(subject: dict, teacher_id: int) -> None:
    if subject["teacher_id"] != teacher_id:
        raise SystemExit(
            f"Доступ запрещён: предмет «{subject['name']}» закреплён за преподавателем "
            f"{subject['teacher_id']}, а не {teacher_id}."
        )


def require_department_subjects(subjects: Iterable[dict], department_id: int) -> None:
    forbidden = [subject for subject in subjects if subject["department_id"] != department_id]
    if forbidden:
        names = ", ".join(subject["name"] for subject in forbidden)
        raise SystemExit(
            f"Доступ запрещён: следующие предметы не относятся к кафедре {department_id}: {names}"
        )


def iter_students(groups: Iterable[dict]):
    for group in groups:
        for student in group["students"]:
            yield group, student


def student_subject(student: dict, subject_id: int) -> dict:
    for subject in student["subjects"]:
        if subject["subject_id"] == subject_id:
            return subject
    raise SystemExit(
        f"Для студента {student['student_ref']} отсутствуют данные предмета {subject_id}."
    )


def has_personal_data_consent(student: dict) -> bool:
    consent = student.get("personal_data_consent")
    if isinstance(consent, dict):
        return bool(consent.get("accepted"))
    return bool(consent)


def short_person_name(value: str) -> str:
    """Convert a full name to the compact report form: Surname I.O."""
    parts = str(value or "").split()
    if len(parts) < 2:
        return " ".join(parts)

    initials = "".join(f"{part[0].upper()}." for part in parts[1:] if part)
    return f"{parts[0]} {initials}"


def student_display_label(student: dict) -> str:
    if has_personal_data_consent(student):
        label = student.get("student_label") or student.get("student_ref", "")
        return short_person_name(label)
    return student.get("student_ref") or student.get("student_label", "")


def attendance_mark(activity: dict) -> str:
    status = activity.get("attendance_status")
    code = activity.get("attendance_code")
    if status == "absent" or code == "О":
        return "Н"
    if status == "sick_with_certificate" or code == "Б":
        return "Б"
    return ""


def score_to_100(activity: dict) -> float | int | None:
    score = activity.get("score")
    max_score = activity.get("max_score")
    if score is None or max_score in (None, 0):
        return None
    value = round(score / max_score * 100, 2)
    return int(value) if value.is_integer() else value


def new_workbook(title: str) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = title
    workbook.properties.creator = "EJournal report prototype"
    workbook.properties.subject = "Ролевой отчёт по учебным метрикам"
    return workbook


def prepare_output(path: str | Path) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    return output


def title_row(sheet, title: str, last_column: int, subtitle: str | None = None) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34

    if subtitle:
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        cell = sheet.cell(2, 1, subtitle)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(size=10, color=MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[2].height = 25


def style_section_header(sheet, row: int, start_column: int, end_column: int, value: str) -> None:
    if end_column < start_column:
        return
    sheet.merge_cells(
        start_row=row,
        start_column=start_column,
        end_row=row,
        end_column=end_column,
    )
    cell = sheet.cell(row, start_column, value)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in range(start_column, end_column + 1):
        sheet.cell(row, column).border = THIN_BORDER


def style_headers(sheet, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=CYAN)
        cell.font = Font(bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[row].height = 42


def style_data_cell(cell, horizontal: str = "center") -> None:
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    cell.font = Font(color=TEXT)


def performance_fill(value: float | int | None) -> PatternFill:
    if value is None:
        return PatternFill("solid", fgColor=WHITE)
    if value >= 85:
        color = GREEN
    elif value >= 60:
        color = YELLOW
    else:
        color = RED
    return PatternFill("solid", fgColor=color)


def add_bar_chart(
    sheet,
    title: str,
    data_column: int,
    category_column: int,
    min_row: int,
    max_row: int,
    anchor: str,
    y_title: str = "Процент",
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.x_axis.title = ""
    chart.height = 8
    chart.width = 15
    data = Reference(sheet, min_col=data_column, min_row=min_row - 1, max_row=max_row)
    categories = Reference(sheet, min_col=category_column, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    sheet.add_chart(chart, anchor)


def set_widths(sheet, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
