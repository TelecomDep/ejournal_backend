from __future__ import annotations

import argparse
from collections import defaultdict
from math import ceil, floor
from pathlib import Path
from statistics import mean

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_common import (
    BLUE,
    BORDER_COLOR,
    CYAN,
    GREEN,
    LIGHT_BLUE,
    MUTED,
    PALE_RED,
    RED,
    TEXT,
    THIN_BORDER,
    WHITE,
    YELLOW,
    has_personal_data_consent,
    iter_students,
    load_metrics,
    new_workbook,
    prepare_output,
    require_department_subjects,
    select_groups,
    select_subjects,
    set_widths,
    student_display_label,
    student_subject,
)


SCRIPT_DIR = Path(__file__).resolve().parent
LOW_SCORE_FILL = "EAC99D"
GRAY_FILL = "C9C9C9"
MARK_COLORS = {
    5: GREEN,
    4: YELLOW,
    3: LOW_SCORE_FILL,
    2: RED,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Сформировать XLSX-рейтинг заведующего кафедрой."
    )
    parser.add_argument("--department-id", type=int, required=True, help="ID кафедры")
    parser.add_argument("--subject-ids", type=int, nargs="*", help="Предметы; без параметра — вся кафедра")
    parser.add_argument("--group-ids", type=int, nargs="*", help="Группы; без параметра — все группы")
    parser.add_argument(
        "--input",
        default=str(SCRIPT_DIR / "fake_metrics.json"),
        help="Путь к исходному JSON",
    )
    parser.add_argument(
        "--output",
        default=str(SCRIPT_DIR / "output" / "department_head_report.xlsx"),
        help="Путь к XLSX",
    )
    return parser.parse_args()


def safe_mean(values: list[float | int]) -> float:
    numeric_values = [value for value in values if isinstance(value, (int, float))]
    return round(mean(numeric_values), 2) if numeric_values else 0


def score_fill(value: float | int | None) -> PatternFill:
    if value is None:
        return PatternFill("solid", fgColor=GRAY_FILL)
    if value >= 80:
        color = GREEN
    elif value >= 60:
        color = YELLOW
    elif value >= 40:
        color = LOW_SCORE_FILL
    else:
        color = PALE_RED
    return PatternFill("solid", fgColor=color)


def mark_fill(mark: int) -> PatternFill:
    return PatternFill("solid", fgColor=MARK_COLORS[mark])


def rating_fill(value: float | int | None) -> PatternFill:
    if value is None:
        return PatternFill("solid", fgColor=GRAY_FILL)
    return mark_fill(rating_mark(value))


def rating_mark(value: float | int) -> int:
    if value >= 80:
        return 5
    if value >= 60:
        return 4
    if value >= 40:
        return 3
    return 2


def axis_bounds(values: list[float | int], padding: float = 1.0) -> tuple[int, int]:
    if not values:
        return 0, 100
    low = max(0, floor(min(values) - padding))
    high = min(100, ceil(max(values) + padding))
    if high - low < 4:
        middle = (high + low) / 2
        low = max(0, floor(middle - 2))
        high = min(100, ceil(middle + 2))
    return low, high


def short_subject_name(subject: dict) -> str:
    return subject.get("short_name") or subject.get("subject_code") or subject["name"]


def build_student_rows(groups: list[dict], subjects: list[dict]) -> list[dict]:
    rows = []
    for group, student in iter_students(groups):
        subject_values = []
        attendance_values = []
        for subject in subjects:
            metrics = next(
                (item for item in student["subjects"] if item["subject_id"] == subject["subject_id"]),
                None,
            )
            subject_values.append(
                metrics["assessment_summary"]["performance_percent"] if metrics else None
            )
            if metrics:
                attendance_values.append(metrics["attendance_summary"]["attendance_percent"])

        available_subject_values = [value for value in subject_values if value is not None]
        if not available_subject_values:
            continue

        rating = safe_mean(available_subject_values)
        attendance = safe_mean(attendance_values)
        rows.append(
            {
                "label": student_display_label(student),
                "has_consent": has_personal_data_consent(student),
                "group": group["group_name"],
                "subjects": subject_values,
                "rating": rating,
                "attendance": attendance,
                "mark": rating_mark(rating),
            }
        )

    return sorted(rows, key=lambda item: (-item["rating"], -item["attendance"], item["group"], item["label"]))


def style_header_cell(cell, fill: str = CYAN) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=TEXT, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_body_cell(cell, horizontal: str = "center") -> None:
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    cell.font = Font(color=TEXT, size=10)


def write_rating_sheet(sheet, department_name: str, semester_title: str, subjects: list[dict], rows: list[dict]) -> dict:
    headers = ["номер", "ФИО / хеш", "Группа"]
    headers.extend(short_subject_name(subject) for subject in subjects)
    headers.extend(["Итоговый\nрейтинг", "Посещаемость"])

    last_column = len(headers)
    header_row = 1
    first_data_row = 2
    subject_start = 4
    rating_column = last_column - 1
    attendance_column = last_column

    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        if column <= 3 or column == rating_column:
            style_header_cell(cell, CYAN)
        else:
            style_header_cell(cell, BLUE)
            cell.font = Font(bold=True, color=WHITE, size=10)

    sheet.row_dimensions[header_row].height = 36

    for row_offset, item in enumerate(rows):
        row = first_data_row + row_offset
        values = [row_offset + 1, item["label"], item["group"]]
        values.extend(item["subjects"])
        values.extend([item["rating"], item["attendance"]])

        row_fill = mark_fill(item["mark"])
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            style_body_cell(cell, "left" if column == 2 else "center")
            if column <= 3:
                cell.fill = row_fill
            elif subject_start <= column < rating_column:
                cell.fill = rating_fill(value)
                cell.number_format = "0.00"
            elif column == rating_column:
                cell.fill = row_fill
                cell.font = Font(bold=True, color=TEXT, size=10)
                cell.number_format = "0.00"
            elif column == attendance_column:
                cell.fill = score_fill(value)
                cell.number_format = "0.00"

            if column == 2 and not item["has_consent"]:
                cell.font = Font(color=MUTED, size=8)

    last_data_row = first_data_row + len(rows) - 1
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_column)}{last_data_row}"
    sheet.sheet_view.showGridLines = False

    sheet.cell(last_data_row + 2, 1, f"{department_name} | {semester_title}")
    sheet.cell(last_data_row + 2, 1).font = Font(italic=True, color=MUTED, size=9)
    sheet.merge_cells(start_row=last_data_row + 2, start_column=1, end_row=last_data_row + 2, end_column=last_column)

    set_widths(sheet, {1: 8, 2: 22, 3: 14, rating_column: 15, attendance_column: 15})
    for column in range(subject_start, rating_column):
        sheet.column_dimensions[get_column_letter(column)].width = 14

    legend_column = last_column + 5
    sheet.cell(1, legend_column + 1, "сред.\nпосещаемость")
    style_header_cell(sheet.cell(1, legend_column + 1), CYAN)
    sheet.column_dimensions[get_column_letter(legend_column)].width = 8
    sheet.column_dimensions[get_column_letter(legend_column + 1)].width = 16

    by_mark = defaultdict(list)
    for item in rows:
        by_mark[item["mark"]].append(item["attendance"])

    for offset, mark in enumerate([5, 4, 3, 2], 2):
        mark_cell = sheet.cell(offset, legend_column, mark)
        attendance = safe_mean(by_mark[mark]) if by_mark[mark] else None
        attendance_cell = sheet.cell(offset, legend_column + 1, attendance)
        for cell in (mark_cell, attendance_cell):
            style_body_cell(cell)
            cell.fill = mark_fill(mark)
        attendance_cell.fill = PatternFill("solid", fgColor=WHITE)
        if attendance is not None:
            attendance_cell.number_format = "0.00"

    return {
        "last_column": last_column,
        "first_data_row": first_data_row,
        "last_data_row": last_data_row,
        "subject_start": subject_start,
        "rating_column": rating_column,
        "attendance_column": attendance_column,
    }


def write_chart_source(charts, subjects: list[dict], rows: list[dict]) -> dict:
    charts.sheet_view.showGridLines = False
    charts.cell(1, 1, "Графики сводки")
    charts.cell(1, 1).fill = PatternFill("solid", fgColor=BLUE)
    charts.cell(1, 1).font = Font(size=16, bold=True, color=WHITE)
    charts.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    charts.row_dimensions[1].height = 30

    subject_header_row = 3
    charts.cell(subject_header_row, 1, "Предмет")
    charts.cell(subject_header_row, 2, "Средний рейтинг")
    for column in range(1, 3):
        style_header_cell(charts.cell(subject_header_row, column))

    for index, subject in enumerate(subjects, subject_header_row + 1):
        subject_position = index - subject_header_row - 1
        values = [item["subjects"][subject_position] for item in rows if item["subjects"][subject_position] is not None]
        charts.cell(index, 1, short_subject_name(subject))
        charts.cell(index, 2, safe_mean(values))
        for column in range(1, 3):
            style_body_cell(charts.cell(index, column), "left" if column == 1 else "center")
            if column == 2:
                charts.cell(index, column).fill = rating_fill(charts.cell(index, column).value)
                charts.cell(index, column).number_format = "0.00"

    group_header_row = 3
    charts.cell(group_header_row, 5, "Группа")
    charts.cell(group_header_row, 6, "Итоговый рейтинг")
    charts.cell(group_header_row, 7, "Посещаемость")
    for column in range(5, 8):
        style_header_cell(charts.cell(group_header_row, column))

    grouped = defaultdict(list)
    for item in rows:
        grouped[item["group"]].append(item)
    group_first_row = group_header_row + 1
    for row_offset, group_name in enumerate(sorted(grouped)):
        sheet_row = group_first_row + row_offset
        items = grouped[group_name]
        charts.cell(sheet_row, 5, group_name)
        charts.cell(sheet_row, 6, safe_mean([item["rating"] for item in items]))
        charts.cell(sheet_row, 7, safe_mean([item["attendance"] for item in items]))
        for column in range(5, 8):
            style_body_cell(charts.cell(sheet_row, column), "left" if column == 5 else "center")
            if column in (6, 7):
                if column == 6:
                    charts.cell(sheet_row, column).fill = rating_fill(charts.cell(sheet_row, column).value)
                else:
                    charts.cell(sheet_row, column).fill = score_fill(charts.cell(sheet_row, column).value)
                charts.cell(sheet_row, column).number_format = "0.00"

    mark_header_row = 3
    charts.cell(mark_header_row, 10, "Оценка")
    charts.cell(mark_header_row, 11, "Студентов")
    for column in range(10, 12):
        style_header_cell(charts.cell(mark_header_row, column))

    mark_counts = defaultdict(int)
    for item in rows:
        mark_counts[item["mark"]] += 1
    visible_marks = [mark for mark in [5, 4, 3, 2] if mark_counts[mark] > 0]
    for offset, mark in enumerate(visible_marks, mark_header_row + 1):
        charts.cell(offset, 10, mark)
        charts.cell(offset, 11, mark_counts[mark])
        for column in range(10, 12):
            style_body_cell(charts.cell(offset, column))
            charts.cell(offset, column).fill = mark_fill(mark)

    set_widths(charts, {1: 18, 2: 16, 5: 14, 6: 17, 7: 15, 10: 10, 11: 12})

    return {
        "subject_first_row": subject_header_row + 1,
        "subject_last_row": subject_header_row + len(subjects),
        "group_first_row": group_first_row,
        "group_last_row": group_first_row + len(grouped) - 1,
        "mark_first_row": mark_header_row + 1,
        "mark_last_row": mark_header_row + len(visible_marks),
        "subject_axis": axis_bounds([charts.cell(row, 2).value for row in range(subject_header_row + 1, subject_header_row + len(subjects) + 1)], 2),
        "group_rating_axis": axis_bounds([charts.cell(row, 6).value for row in range(group_first_row, group_first_row + len(grouped))], 0.5),
        "group_attendance_axis": axis_bounds([charts.cell(row, 7).value for row in range(group_first_row, group_first_row + len(grouped))], 0.5),
    }


def add_chart(
    sheet,
    title: str,
    data_column: int,
    category_column: int,
    first_row: int,
    last_row: int,
    anchor: str,
    chart_type: str = "col",
    min_y: int = 0,
    max_y: int | None = 100,
) -> None:
    chart = BarChart()
    chart.type = chart_type
    chart.style = 10
    chart.title = title
    chart.height = 8
    chart.width = 14

    if chart_type == "bar":
        chart.x_axis.scaling.min = min_y
        if max_y is not None:
            chart.x_axis.scaling.max = max_y
    else:
        chart.y_axis.scaling.min = min_y
        if max_y is not None:
            chart.y_axis.scaling.max = max_y

    data = Reference(sheet, min_col=data_column, min_row=first_row - 1, max_row=last_row)
    categories = Reference(sheet, min_col=category_column, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    if chart.series:
        chart.series[0].tx = SeriesLabel(v=str(sheet.cell(first_row - 1, data_column).value))
        chart.series[0].graphicalProperties.solidFill = "5B7FBD"
        chart.series[0].graphicalProperties.line.solidFill = "5B7FBD"
    chart.legend = None
    sheet.add_chart(chart, anchor)


def build_report(args: argparse.Namespace) -> str:
    data = load_metrics(args.input)
    groups = select_groups(data, args.group_ids)

    if args.subject_ids:
        subjects = select_subjects(data, args.subject_ids)
    else:
        subjects = [
            subject for subject in data["subjects"] if subject["department_id"] == args.department_id
        ]
        if not subjects:
            raise SystemExit(f"Для кафедры {args.department_id} не найдено предметов.")
    require_department_subjects(subjects, args.department_id)

    department = next(
        (item for item in data.get("departments", []) if item["department_id"] == args.department_id),
        None,
    )
    department_name = department["department_name"] if department else f"Кафедра {args.department_id}"
    semester_title = data.get("semester", {}).get("title", "семестр не указан")

    rows = build_student_rows(groups, subjects)
    workbook = new_workbook(f"Рейтинг кафедры: {department_name}")
    rating = workbook.create_sheet("Статистика")
    charts = workbook.create_sheet("Графики")

    write_rating_sheet(rating, department_name, semester_title, subjects, rows)
    chart_ranges = write_chart_source(charts, subjects, rows)

    add_chart(
        charts,
        "Средний рейтинг по предметам",
        2,
        1,
        chart_ranges["subject_first_row"],
        chart_ranges["subject_last_row"],
        "A12",
        chart_type="bar",
        min_y=chart_ranges["subject_axis"][0],
        max_y=chart_ranges["subject_axis"][1],
    )
    add_chart(
        charts,
        "Итоговый рейтинг по группам",
        6,
        5,
        chart_ranges["group_first_row"],
        chart_ranges["group_last_row"],
        "J12",
        min_y=chart_ranges["group_rating_axis"][0],
        max_y=chart_ranges["group_rating_axis"][1],
    )
    add_chart(
        charts,
        "Посещаемость по группам",
        7,
        5,
        chart_ranges["group_first_row"],
        chart_ranges["group_last_row"],
        "J29",
        min_y=chart_ranges["group_attendance_axis"][0],
        max_y=chart_ranges["group_attendance_axis"][1],
    )
    add_chart(
        charts,
        "Распределение оценок",
        11,
        10,
        chart_ranges["mark_first_row"],
        chart_ranges["mark_last_row"],
        "A29",
        max_y=None,
    )

    output = prepare_output(args.output)
    workbook.save(output)
    return str(output)


if __name__ == "__main__":
    arguments = parse_args()
    path = build_report(arguments)
    print(f"Отчёт заведующего кафедрой создан: {path}")
